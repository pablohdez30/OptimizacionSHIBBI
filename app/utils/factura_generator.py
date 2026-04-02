import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XlImage
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_RIGHT

from app.utils.output_path import get_output_path


def generar_factura_excel(app, factura_id):
    """Generate Excel invoice using PresupuestoPlantilla.xlsx as base template."""
    factura = app.factura_model.obtener(factura_id)
    if not factura:
        return None

    # Find template
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    template_path = os.path.join(base_dir, "PresupuestoPlantilla.xlsx")

    if not os.path.exists(template_path):
        return None

    # Config
    empresa_nombre = app.config_model.obtener("empresa_nombre") or "CAESPAN ARGUMENT S.L."
    empresa_dir = app.config_model.obtener("empresa_direccion") or ""
    empresa_ciudad = app.config_model.obtener("empresa_ciudad") or ""
    empresa_cif = app.config_model.obtener("empresa_cif") or ""
    empresa_cuenta = app.config_model.obtener("empresa_cuenta_bancaria") or ""
    iva_pct = app.config_model.obtener_float("iva_porcentaje", 21)

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active

    # Add logo image
    logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              "assets", "logo_shibbi.jpg")
    if os.path.exists(logo_path):
        logo = XlImage(logo_path)
        logo.width = 180
        logo.height = 30
        ws.add_image(logo, "A1")

    # Fill company data
    ws["E1"] = empresa_nombre
    ws["E2"] = empresa_dir
    ws["E3"] = empresa_ciudad
    ws["E4"] = empresa_cif

    # Change header to FACTURA
    ws["A3"] = "FACTURA Nº:"
    ws["A3"].font = Font(bold=True, size=12)
    ws["B3"] = factura["numero_factura"]
    ws["B3"].font = Font(bold=True, size=12)

    try:
        fecha_obj = datetime.strptime(factura["fecha"], "%Y-%m-%d")
        ws["B4"] = fecha_obj
        ws["B4"].number_format = "DD/MM/YYYY"
    except (ValueError, TypeError):
        ws["B4"] = factura["fecha"] or ""

    # Client
    ws["B5"] = factura["cliente_nombre"]
    ws["B5"].font = Font(bold=True, size=12)
    ws["B6"] = factura["cliente_direccion"] or ""
    ws["B8"] = factura["cliente_nif"] or ""

    # Clear product rows (10-19)
    for r in range(10, 20):
        for c in [1, 5, 6, 7]:
            ws.cell(row=r, column=c).value = None

    # Set column headers for factura format
    ws.cell(row=9, column=5, value="Unidades").font = Font(bold=True, size=10)
    ws.cell(row=9, column=6, value="PX").font = Font(bold=True, size=10)
    ws.cell(row=9, column=7, value="Total").font = Font(bold=True, size=10)

    # Fill lines
    lineas = app.factura_model.obtener_lineas(factura_id)
    row = 10
    total_base = 0
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    money_fmt = '#,##0.00 €'

    for linea in lineas:
        line_total = linea["unidades"] * linea["precio_unitario"]
        total_base += line_total

        concepto = linea["concepto"]
        if linea["es_porte"]:
            concepto = f"(Porte) {concepto}"

        ws.cell(row=row, column=1, value=concepto).font = bold_font
        ws.cell(row=row, column=5, value=linea["unidades"]).font = normal_font
        ws.cell(row=row, column=6, value=linea["precio_unitario"]).font = normal_font
        ws.cell(row=row, column=6).number_format = money_fmt
        ws.cell(row=row, column=7, value=line_total).font = normal_font
        ws.cell(row=row, column=7).number_format = money_fmt
        row += 1

        if row >= 18:
            break

    # Adelanto line (negative)
    adelanto = factura["adelanto_importe"] or 0
    if adelanto > 0 and row < 18:
        desc = factura["adelanto_descripcion"] or "Adelanto 50%"
        ws.cell(row=row, column=1, value=desc).font = bold_font
        ws.cell(row=row, column=7, value=-adelanto).font = Font(bold=True, size=11, color="FF0000")
        ws.cell(row=row, column=7).number_format = money_fmt
        total_base -= adelanto
        row += 1

    # Totals
    iva = total_base * (iva_pct / 100)
    total_final = total_base + iva

    ws["D21"] = "Total Base"
    ws["G21"] = total_base
    ws["G21"].number_format = money_fmt

    ws["D22"] = f"I.V.A {iva_pct:.0f}%"
    ws["G22"] = iva
    ws["G22"].number_format = money_fmt

    ws["D23"] = "TOTAL"
    ws["D23"].font = bold_font
    ws["G23"] = total_final
    ws["G23"].number_format = money_fmt
    ws["G23"].font = bold_font

    # NO condiciones de pago section - clear those rows
    for r in range(24, 27):
        ws.cell(row=r, column=1).value = None

    # Only bank transfer line
    ws["A24"] = f"Pago mediante transferencia bancaria: CC: {empresa_cuenta}"
    ws["A24"].font = Font(bold=True, size=12)

    # Save
    filepath = get_output_path("FACTURAS", factura["numero_factura"],
                               factura["cliente_nombre"], "xlsx", app=app)
    try:
        wb.save(filepath)
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        filepath = filepath.replace(".xlsx", f"_{ts}.xlsx")
        wb.save(filepath)
    return filepath


def generar_factura_pdf(app, factura_id):
    """Generate PDF invoice matching company template style."""
    factura = app.factura_model.obtener(factura_id)
    if not factura:
        return None

    empresa_nombre = app.config_model.obtener("empresa_nombre") or "CAESPAN ARGUMENT S.L."
    empresa_dir = app.config_model.obtener("empresa_direccion") or ""
    empresa_ciudad = app.config_model.obtener("empresa_ciudad") or ""
    empresa_cif = app.config_model.obtener("empresa_cif") or ""
    empresa_cuenta = app.config_model.obtener("empresa_cuenta_bancaria") or ""
    iva_pct = app.config_model.obtener_float("iva_porcentaje", 21)

    # Output path
    filepath = get_output_path("FACTURAS", factura["numero_factura"],
                               factura["cliente_nombre"], "pdf", app=app)

    # Colors
    black = HexColor("#000000")
    border_color = HexColor("#000000")

    # Styles
    s_normal = ParagraphStyle("N", fontSize=11, fontName="Helvetica", textColor=black, leading=14)
    s_bold = ParagraphStyle("B", fontSize=11, fontName="Helvetica-Bold", textColor=black, leading=14)
    s_bold12 = ParagraphStyle("B12", fontSize=12, fontName="Helvetica-Bold", textColor=black, leading=15)
    s_right = ParagraphStyle("R", fontSize=11, fontName="Helvetica", textColor=black, alignment=TA_RIGHT, leading=14)
    s_right_bold = ParagraphStyle("RB", fontSize=11, fontName="Helvetica-Bold", textColor=black, alignment=TA_RIGHT, leading=14)
    s_company = ParagraphStyle("C", fontSize=11, fontName="Helvetica-Bold", textColor=black, alignment=TA_RIGHT, leading=15)

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            leftMargin=20 * mm, rightMargin=20 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    elements = []

    # --- LOGO IMAGE ---
    logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              "assets", "logo_shibbi.jpg")
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=50 * mm, height=14 * mm)
        header_data = [[logo, Paragraph(
            f"{empresa_nombre}<br/>{empresa_dir}<br/>{empresa_ciudad}<br/>{empresa_cif}",
            s_company)]]
    else:
        s_logo = ParagraphStyle("Logo", fontSize=32, fontName="Helvetica-Bold",
                                 textColor=HexColor("#7FAEAB"), leading=36)
        header_data = [[Paragraph("SHIBBI", s_logo), Paragraph(
            f"{empresa_nombre}<br/>{empresa_dir}<br/>{empresa_ciudad}<br/>{empresa_cif}",
            s_company)]]

    ht = Table(header_data, colWidths=[90 * mm, 80 * mm])
    ht.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    elements.append(ht)

    # --- Invoice info + Client info ---
    try:
        fecha_str = datetime.strptime(factura["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        fecha_str = factura["fecha"] or ""

    info_data = [
        [Paragraph("<b>FACTURA  Nº:</b>", s_bold), Paragraph(factura["numero_factura"], s_bold), "", ""],
        [Paragraph("FECHA:", s_normal), Paragraph(fecha_str, s_normal), "", ""],
        [Paragraph("CLIENTE:", s_normal), Paragraph(f"<b>{factura['cliente_nombre']}</b>", s_bold12), "", ""],
        [Paragraph("DIRECCIÓN:", s_normal), Paragraph(factura["cliente_direccion"] or "", s_normal), "", ""],
        [Paragraph("C.I.F./N.I.F:", s_normal), "", "", ""],
        ["", Paragraph(factura["cliente_nif"] or "", s_normal), "", ""],
    ]

    it = Table(info_data, colWidths=[35 * mm, 55 * mm, 40 * mm, 40 * mm])
    it.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("LINEBELOW", (0, 0), (1, 0), 1, border_color),
    ]))
    elements.append(it)
    elements.append(Spacer(1, 8 * mm))

    # --- Column headers bar ---
    bar_data = [[
        Paragraph("<b>Concepto</b>", s_bold),
        Paragraph("<b>Unidades</b>", s_right_bold),
        Paragraph("<b>PX</b>", s_right_bold),
        Paragraph("<b>Total</b>", s_right_bold),
    ]]
    bt = Table(bar_data, colWidths=[85 * mm, 25 * mm, 30 * mm, 30 * mm])
    bt.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 2, border_color),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, border_color),
        ("TOPPADDING", (0, 0), (-1, 0), 4),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
    ]))
    elements.append(bt)

    # --- Lines ---
    lineas = app.factura_model.obtener_lineas(factura_id)
    total_base = 0

    for linea in lineas:
        line_total = linea["unidades"] * linea["precio_unitario"]
        total_base += line_total

        concepto = linea["concepto"]
        if linea["es_porte"]:
            concepto = f"(Porte) {concepto}"

        prod_row = [[
            Paragraph(f"<b>{concepto}</b>", s_bold),
            Paragraph(f"{linea['unidades']:g}", s_right),
            Paragraph(f"{linea['precio_unitario']:,.2f}", s_right),
            Paragraph(f"{line_total:,.2f}", s_right),
        ]]
        pt = Table(prod_row, colWidths=[85 * mm, 25 * mm, 30 * mm, 30 * mm])
        pt.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, 0), 3),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 1),
        ]))
        elements.append(pt)
        elements.append(Spacer(1, 1 * mm))

    # Adelanto (negative)
    adelanto = factura["adelanto_importe"] or 0
    if adelanto > 0:
        desc = factura["adelanto_descripcion"] or "Adelanto 50%"
        adel_row = [[
            Paragraph(f"<b>{desc}</b>", s_bold),
            Paragraph("", s_right),
            Paragraph("", s_right),
            Paragraph(f"-{adelanto:,.2f}", s_right_bold),
        ]]
        at = Table(adel_row, colWidths=[85 * mm, 25 * mm, 30 * mm, 30 * mm])
        at.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, 0), 3),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 1),
        ]))
        elements.append(at)
        total_base -= adelanto

    elements.append(Spacer(1, 10 * mm))

    # --- Totals box ---
    iva = total_base * (iva_pct / 100)
    total_final = total_base + iva

    totals_data = [
        [Paragraph("Total Base", s_normal), Paragraph(f"{total_base:,.2f} €", s_right)],
        [Paragraph(f"I.V.A {iva_pct:.0f}%", s_normal), Paragraph(f"{iva:,.2f} €", s_right)],
        [Paragraph("<b>TOTAL</b>", s_bold), Paragraph(f"<b>{total_final:,.2f} €</b>", s_right_bold)],
    ]
    tt = Table(totals_data, colWidths=[40 * mm, 40 * mm])
    tt.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 1, border_color),
        ("LINEABOVE", (0, 2), (-1, 2), 0.5, border_color),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))

    tw = Table([[None, tt]], colWidths=[90 * mm, 80 * mm])
    tw.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements.append(tw)
    elements.append(Spacer(1, 5 * mm))

    # NO condiciones de pago - only bank transfer line
    elements.append(Paragraph(
        f"<b>Pago mediante transferencia bancaria: CC: {empresa_cuenta}</b>", s_bold))

    try:
        doc.build(elements)
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        filepath = filepath.replace(".pdf", f"_{ts}.pdf")
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                leftMargin=20 * mm, rightMargin=20 * mm,
                                topMargin=15 * mm, bottomMargin=15 * mm)
        doc.build(elements)

    return filepath
