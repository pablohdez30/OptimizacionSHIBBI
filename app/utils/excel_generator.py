import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def generar_excel_presupuesto(app, presupuesto_id):
    """Generate an internal Excel with full cost breakdown for a budget."""
    pres = app.presupuesto_model.obtener(presupuesto_id)
    if not pres:
        return None

    wb = Workbook()

    # --- Sheet 1: Cost breakdown (internal) ---
    ws = wb.active
    ws.title = "Desglose Costes"

    # Styles
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    subheader_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="e94560", end_color="e94560", fill_type="solid")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    money_format = '#,##0.00€'
    thin_border = Border(
        left=Side(style="thin", color="d0d0d0"),
        right=Side(style="thin", color="d0d0d0"),
        top=Side(style="thin", color="d0d0d0"),
        bottom=Side(style="thin", color="d0d0d0"),
    )
    light_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=f"DESGLOSE DE COSTES - {pres['numero_presupuesto']}")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    # Client info
    ws.cell(row=row, column=1, value="Cliente:").font = bold_font
    ws.cell(row=row, column=2, value=pres["cliente_nombre"]).font = normal_font
    ws.cell(row=row, column=4, value="Fecha:").font = bold_font
    ws.cell(row=row, column=5, value=pres["fecha"]).font = normal_font
    row += 2

    # Lines
    lineas = app.presupuesto_model.obtener_lineas(presupuesto_id)
    gran_total_coste = 0
    gran_total_cliente = 0

    for linea in lineas:
        # Product header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1,
                       value=f"{linea['nombre_producto']} — {linea['descripcion']}")
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 25
        row += 1

        # Column headers
        headers = ["Categoría", "Proveedor", "Descripción", "Cantidad",
                    "Precio Ud.", "Total"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = Font(name="Calibri", size=10, bold=True, color="666666")
            cell.fill = light_fill
            cell.border = thin_border
        row += 1

        # Detail rows
        detalles = app.presupuesto_model.obtener_detalles_coste(linea["id"])
        coste_linea = 0

        for det in detalles:
            ws.cell(row=row, column=1, value=det["categoria_nombre"] or "").font = normal_font
            ws.cell(row=row, column=2, value=det["proveedor_nombre"] or "").font = normal_font
            ws.cell(row=row, column=3, value=det["descripcion"] or "").font = normal_font
            ws.cell(row=row, column=4, value=det["cantidad"]).font = normal_font
            ws.cell(row=row, column=5, value=det["precio_unitario"]).font = normal_font
            ws.cell(row=row, column=5).number_format = money_format
            ws.cell(row=row, column=6, value=det["precio_total"]).font = normal_font
            ws.cell(row=row, column=6).number_format = money_format

            for c in range(1, 7):
                ws.cell(row=row, column=c).border = thin_border

            coste_linea += det["precio_total"]
            row += 1

        # Line totals
        precio_cliente = linea["precio_unitario_final"]
        ws.cell(row=row, column=4, value="COSTE:").font = bold_font
        ws.cell(row=row, column=5).font = bold_font
        ws.cell(row=row, column=6, value=coste_linea).font = bold_font
        ws.cell(row=row, column=6).number_format = money_format
        row += 1

        ws.cell(row=row, column=4, value=f"Margen {linea['margen_porcentaje']:.0f}%:").font = bold_font
        ws.cell(row=row, column=6, value=precio_cliente).font = bold_font
        ws.cell(row=row, column=6).number_format = money_format

        ws.cell(row=row, column=7, value=f"x{linea['cantidad']}").font = normal_font
        row += 1

        gran_total_coste += coste_linea * linea["cantidad"]
        gran_total_cliente += precio_cliente * linea["cantidad"]
        row += 1

    # Grand total
    row += 1
    ws.cell(row=row, column=4, value="COSTE TOTAL:").font = bold_font
    ws.cell(row=row, column=6, value=gran_total_coste).font = bold_font
    ws.cell(row=row, column=6).number_format = money_format
    row += 1

    ws.cell(row=row, column=4, value="PRECIO CLIENTE:").font = bold_font
    ws.cell(row=row, column=6, value=gran_total_cliente).font = Font(name="Calibri", size=12, bold=True, color="e94560")
    ws.cell(row=row, column=6).number_format = money_format
    row += 1

    iva_pct = app.config_model.obtener_float("iva_porcentaje", 21)
    iva = gran_total_cliente * (iva_pct / 100)
    ws.cell(row=row, column=4, value=f"IVA {iva_pct:.0f}%:").font = bold_font
    ws.cell(row=row, column=6, value=iva).font = bold_font
    ws.cell(row=row, column=6).number_format = money_format
    row += 1

    total_final = gran_total_cliente + iva
    ws.cell(row=row, column=4, value="TOTAL CON IVA:").font = Font(name="Calibri", size=13, bold=True)
    ws.cell(row=row, column=6, value=total_final).font = Font(name="Calibri", size=13, bold=True, color="e94560")
    ws.cell(row=row, column=6).number_format = money_format
    row += 1

    beneficio = gran_total_cliente - gran_total_coste
    ws.cell(row=row, column=4, value="BENEFICIO:").font = bold_font
    ws.cell(row=row, column=6, value=beneficio).font = Font(name="Calibri", size=11, bold=True, color="2d6a4f")
    ws.cell(row=row, column=6).number_format = money_format

    # Save in client folder
    base_output = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                "output")
    # Create folder: output/ClienteName/
    safe_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_"
                        for c in pres["cliente_nombre"]).strip()
    output_dir = os.path.join(base_output, safe_name)
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Desglose_{pres['numero_presupuesto']}_{safe_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
