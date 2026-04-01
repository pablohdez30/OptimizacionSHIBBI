import os
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers


def generar_pdf_presupuesto(app, presupuesto_id):
    """Generate presupuesto as Excel file matching the original company template.
    Uses PresupuestoPlantilla.xlsx as the base template."""
    pres = app.presupuesto_model.obtener(presupuesto_id)
    if not pres:
        return None

    # Find template
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    template_path = os.path.join(base_dir, "PresupuestoPlantilla.xlsx")

    if not os.path.exists(template_path):
        # Fallback: generate from scratch if template missing
        return _generar_pdf_fallback(app, presupuesto_id)

    # Config
    empresa_nombre = app.config_model.obtener("empresa_nombre") or "CAESPAN ARGUMENT S.L."
    empresa_dir = app.config_model.obtener("empresa_direccion") or ""
    empresa_ciudad = app.config_model.obtener("empresa_ciudad") or ""
    empresa_cif = app.config_model.obtener("empresa_cif") or ""
    empresa_cuenta = app.config_model.obtener("empresa_cuenta_bancaria") or ""
    iva_pct = app.config_model.obtener_float("iva_porcentaje", 21)

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active

    # Fill in header data
    ws["E1"] = empresa_nombre
    ws["E2"] = empresa_dir
    ws["E3"] = empresa_ciudad
    ws["E4"] = empresa_cif

    # Budget number and date
    ws["B3"] = pres["numero_presupuesto"]
    try:
        fecha_obj = datetime.strptime(pres["fecha"], "%Y-%m-%d")
        ws["B4"] = fecha_obj
        ws["B4"].number_format = "DD/MM/YYYY"
    except (ValueError, TypeError):
        ws["B4"] = pres["fecha"] or ""

    # Client data
    ws["B5"] = pres["cliente_nombre"]
    ws["B5"].font = Font(bold=True, size=12)
    # Direction in row 6, NIF in row 7-8
    ws["B6"] = pres["cliente_direccion"] or ""
    ws["B8"] = pres["cliente_nif"] or ""

    # Clear existing product rows (rows 10-19)
    for r in range(10, 20):
        ws.cell(row=r, column=1).value = None
        ws.cell(row=r, column=7).value = None

    # Fill products
    lineas = app.presupuesto_model.obtener_lineas(presupuesto_id)
    row = 10
    total_base = 0
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    money_fmt = '#,##0.00 €'

    for linea in lineas:
        precio_total = linea["precio_unitario_final"] * linea["cantidad"]
        total_base += precio_total

        # Product name (bold)
        ws.cell(row=row, column=1, value=linea["nombre_producto"]).font = bold_font
        ws.cell(row=row, column=7, value=precio_total).font = normal_font
        ws.cell(row=row, column=7).number_format = money_fmt
        row += 1

        # Description lines
        if linea["descripcion"]:
            for desc_line in linea["descripcion"].split("\n"):
                if desc_line.strip():
                    ws.cell(row=row, column=1, value=desc_line.strip()).font = normal_font
                    row += 1

        # Quantity note if > 1
        if linea["cantidad"] > 1:
            ws.cell(row=row, column=1, value=f"({linea['cantidad']} unidades)").font = normal_font
            row += 1

    # Porte note
    if not pres["incluye_instalacion"]:
        ws.cell(row=18, column=1, value="Porte No incluido").font = bold_font
    else:
        ws.cell(row=18, column=1, value="Porte e instalación incluidos").font = bold_font

    # Totals
    iva = total_base * (iva_pct / 100)
    total_final = total_base + iva

    ws["D21"] = "Total Base"
    ws["G21"] = total_base
    ws["G21"].number_format = money_fmt

    ws["D22"] = f"I.V.A {iva_pct:.0f}%"
    ws["G22"] = iva
    ws["G22"].number_format = money_fmt

    ws["D23"] = "TOTAL"
    ws["D23"].font = bold_font
    ws["G23"] = total_final
    ws["G23"].number_format = money_fmt
    ws["G23"].font = bold_font

    # Payment conditions
    ws["A24"] = "CONDICIONES DE PAGO:"
    ws["A24"].font = Font(bold=True, size=12)
    ws["A25"] = pres["condiciones_pago"] or "50% adelanto-50% antes de la entrega del trabajo."
    ws["A25"].font = Font(size=12)

    try:
        fecha_obj = datetime.strptime(pres["fecha"], "%Y-%m-%d")
        dias = pres["dias_validez"] or 15
        ws["A26"] = f"Los presupuestos caducan a los {dias} días"
    except (ValueError, TypeError):
        ws["A26"] = f"Los presupuestos caducan a los {pres['dias_validez'] or 15} días"
    ws["A26"].font = Font(size=12)

    ws["A27"] = f"Pago mediante transferencia bancaria: CC: {empresa_cuenta}"
    ws["A27"].font = Font(bold=True, size=12)

    # Save
    base_output = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output")
    safe_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_"
                        for c in pres["cliente_nombre"]).strip()
    output_dir = os.path.join(base_output, safe_name)
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Presupuesto_{pres['numero_presupuesto']}_{safe_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def _generar_pdf_fallback(app, presupuesto_id):
    """Fallback: generate PDF if template Excel is missing."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT

    pres = app.presupuesto_model.obtener(presupuesto_id)
    if not pres:
        return None

    empresa_nombre = app.config_model.obtener("empresa_nombre") or "CAESPAN ARGUMENT S.L."
    empresa_dir = app.config_model.obtener("empresa_direccion") or ""
    empresa_ciudad = app.config_model.obtener("empresa_ciudad") or ""
    empresa_cif = app.config_model.obtener("empresa_cif") or ""
    empresa_cuenta = app.config_model.obtener("empresa_cuenta_bancaria") or ""
    iva_pct = app.config_model.obtener_float("iva_porcentaje", 21)

    base_output = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output")
    safe_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_"
                        for c in pres["cliente_nombre"]).strip()
    output_dir = os.path.join(base_output, safe_name)
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Presupuesto_{pres['numero_presupuesto']}_{safe_name}.pdf"
    filepath = os.path.join(output_dir, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []
    dark = HexColor("#1a1a2e")
    accent = HexColor("#e94560")
    white = HexColor("#ffffff")
    tc = HexColor("#333333")

    ts = ParagraphStyle("t", parent=styles["Normal"], fontSize=20, textColor=accent, fontName="Helvetica-Bold")
    cs = ParagraphStyle("c", parent=styles["Normal"], fontSize=10, textColor=tc, fontName="Helvetica", leading=14)

    elements.append(Table([[
        Paragraph(f"PRESUPUESTO Nº: {pres['numero_presupuesto']}", ts),
        Paragraph(f"<b>{empresa_nombre}</b><br/>{empresa_dir}<br/>{empresa_ciudad}<br/>{empresa_cif}", cs)
    ]], colWidths=[95*mm, 75*mm]))
    elements.append(Spacer(1, 8*mm))

    try: fs = datetime.strptime(pres["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y")
    except: fs = pres["fecha"] or ""
    ls = ParagraphStyle("l", parent=styles["Normal"], fontSize=10, textColor=HexColor("#666"), fontName="Helvetica")
    vs = ParagraphStyle("v", parent=styles["Normal"], fontSize=11, textColor=tc, fontName="Helvetica-Bold")
    elements.append(Table([
        [Paragraph("FECHA:", ls), Paragraph(fs, vs), Paragraph("CLIENTE:", ls), Paragraph(pres["cliente_nombre"], vs)],
        ["", "", Paragraph("DIRECCIÓN:", ls), Paragraph(pres["cliente_direccion"] or "", cs)],
        ["", "", Paragraph("C.I.F./N.I.F:", ls), Paragraph(pres["cliente_nif"] or "", cs)],
    ], colWidths=[22*mm, 55*mm, 27*mm, 66*mm]))
    elements.append(Spacer(1, 10*mm))

    lineas = app.presupuesto_model.obtener_lineas(presupuesto_id)
    pd = [[Paragraph("<b>Concepto</b>", ParagraphStyle("", parent=styles["Normal"], fontSize=10, textColor=white, fontName="Helvetica-Bold")),
           Paragraph("<b>Base</b>", ParagraphStyle("", parent=styles["Normal"], fontSize=10, textColor=white, fontName="Helvetica-Bold", alignment=TA_RIGHT))]]
    tb = 0
    for l in lineas:
        pt = l["precio_unitario_final"] * l["cantidad"]; tb += pt
        dp = [l["nombre_producto"]]
        if l["descripcion"]: dp.append(l["descripcion"])
        if l["cantidad"] > 1: dp.append(f"({l['cantidad']} unidades)")
        pd.append([Paragraph("<br/>".join(dp), ParagraphStyle("", parent=styles["Normal"], fontSize=10, textColor=tc, fontName="Helvetica", leading=14)),
                   Paragraph(f"{pt:,.2f}€", ParagraphStyle("", parent=styles["Normal"], fontSize=11, textColor=tc, fontName="Helvetica-Bold", alignment=TA_RIGHT))])
    for _ in range(max(0, 8-len(lineas))): pd.append(["", ""])
    pt = Table(pd, colWidths=[130*mm, 40*mm])
    pt.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),dark),("TEXTCOLOR",(0,0),(-1,0),white),
        ("VALIGN",(0,0),(-1,-1),"TOP"),("TOPPADDING",(0,1),(-1,-1),6),("BOTTOMPADDING",(0,1),(-1,-1),6),
        ("LINEBELOW",(0,0),(-1,-2),0.5,HexColor("#e0e0e0")),("ALIGN",(1,0),(1,-1),"RIGHT"),
        ("BOX",(0,0),(-1,-1),0.5,HexColor("#ccc"))]))
    elements.append(pt)
    if not pres["incluye_instalacion"]:
        elements.append(Spacer(1,3*mm))
        elements.append(Paragraph("Porte No incluido", ParagraphStyle("s", parent=styles["Normal"], fontSize=9, textColor=HexColor("#888"), fontName="Helvetica")))
    elements.append(Spacer(1,8*mm))
    iv = tb*(iva_pct/100); tf = tb+iv
    tt = Table([["Total Base",f"{tb:,.2f}€"],[f"I.V.A {iva_pct:.0f}%",f"{iv:,.2f}€"],["TOTAL",f"{tf:,.2f}€"]], colWidths=[30*mm,40*mm])
    tt.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"RIGHT"),("FONTNAME",(0,0),(-1,-1),"Helvetica"),("FONTSIZE",(0,0),(-1,-1),11),
        ("LINEBELOW",(0,1),(-1,1),0.5,HexColor("#ccc")),("FONTNAME",(0,2),(-1,2),"Helvetica-Bold"),("FONTSIZE",(0,2),(-1,2),13),("TEXTCOLOR",(1,2),(1,2),accent)]))
    elements.append(Table([[None,tt]], colWidths=[100*mm,70*mm]))
    elements.append(Spacer(1,12*mm))
    cb = ParagraphStyle("cb", parent=styles["Normal"], fontSize=9, textColor=tc, fontName="Helvetica-Bold", leading=13)
    cn = ParagraphStyle("cn", parent=styles["Normal"], fontSize=9, textColor=HexColor("#555"), fontName="Helvetica", leading=13)
    elements.append(Paragraph("CONDICIONES DE PAGO:", cb))
    elements.append(Paragraph(pres["condiciones_pago"] or "", cn))
    elements.append(Spacer(1,3*mm))
    elements.append(Paragraph(f"Los presupuestos caducan a los {pres['dias_validez'] or 15} días", cn))
    elements.append(Spacer(1,3*mm))
    elements.append(Paragraph(f"Pago mediante transferencia bancaria: CC: {empresa_cuenta}", cn))
    doc.build(elements)
    return filepath
