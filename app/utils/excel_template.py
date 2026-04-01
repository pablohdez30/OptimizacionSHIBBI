import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XlImage


def generar_excel_plantilla(app, presupuesto_id):
    """Generate Excel budget using the original PresupuestoPlantilla.xlsx template.
    This preserves the company logo (EMF), formatting, and layout."""
    pres = app.presupuesto_model.obtener(presupuesto_id)
    if not pres:
        return None

    # Find template
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    template_path = os.path.join(base_dir, "PresupuestoPlantilla.xlsx")

    if not os.path.exists(template_path):
        return None

    # Config
    empresa_nombre = app.config_model.obtener("empresa_nombre") or "CAESPAN ARGUMENT S.L."
    empresa_dir = app.config_model.obtener("empresa_direccion") or ""
    empresa_ciudad = app.config_model.obtener("empresa_ciudad") or ""
    empresa_cif = app.config_model.obtener("empresa_cif") or ""
    empresa_cuenta = app.config_model.obtener("empresa_cuenta_bancaria") or ""
    iva_pct = app.config_model.obtener_float("iva_porcentaje", 21)

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active

    # Add logo image (original EMF is dropped by openpyxl, so we add JPG)
    logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              "assets", "logo_shibbi.jpg")
    if os.path.exists(logo_path):
        logo = XlImage(logo_path)
        # Scale to match original template size (~4cm wide in the top-left)
        logo.width = 180
        logo.height = 30
        ws.add_image(logo, "A1")

    # Fill company data
    ws["E1"] = empresa_nombre
    ws["E2"] = empresa_dir
    ws["E3"] = empresa_ciudad
    ws["E4"] = empresa_cif

    # Budget number and date
    ws["B3"] = pres["numero_presupuesto"]
    try:
        fecha_obj = datetime.strptime(pres["fecha"], "%Y-%m-%d")
        ws["B4"] = fecha_obj
        ws["B4"].number_format = "DD/MM/YYYY"
    except (ValueError, TypeError):
        ws["B4"] = pres["fecha"] or ""

    # Client
    ws["B5"] = pres["cliente_nombre"]
    ws["B5"].font = Font(bold=True, size=12)
    ws["B6"] = pres["cliente_direccion"] or ""
    ws["B8"] = pres["cliente_nif"] or ""

    # Clear product rows (10-19)
    for r in range(10, 20):
        for c in [1, 7]:
            ws.cell(row=r, column=c).value = None

    # Fill products
    lineas = app.presupuesto_model.obtener_lineas(presupuesto_id)
    row = 10
    total_base = 0
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    money_fmt = '#,##0.00 €'

    for linea in lineas:
        precio_total = linea["precio_unitario_final"] * linea["cantidad"]
        total_base += precio_total

        ws.cell(row=row, column=1, value=linea["nombre_producto"]).font = bold_font
        ws.cell(row=row, column=7, value=precio_total).font = normal_font
        ws.cell(row=row, column=7).number_format = money_fmt
        row += 1

        if linea["descripcion"]:
            for desc_line in linea["descripcion"].split("\n"):
                if desc_line.strip() and row < 18:
                    ws.cell(row=row, column=1, value=desc_line.strip()).font = normal_font
                    row += 1

        if linea["cantidad"] > 1 and row < 18:
            ws.cell(row=row, column=1, value=f"({linea['cantidad']} unidades)").font = normal_font
            row += 1

    # Porte
    if not pres["incluye_instalacion"]:
        ws.cell(row=18, column=1, value="Porte No incluido").font = bold_font
    else:
        ws.cell(row=18, column=1, value="Porte e instalación incluidos").font = bold_font

    # Totals
    iva = total_base * (iva_pct / 100)
    total_final = total_base + iva

    ws["D21"] = "Total Base"
    ws["G21"] = total_base
    ws["G21"].number_format = money_fmt

    ws["D22"] = f"I.V.A {iva_pct:.0f}%"
    ws["G22"] = iva
    ws["G22"].number_format = money_fmt

    ws["D23"] = "TOTAL"
    ws["D23"].font = bold_font
    ws["G23"] = total_final
    ws["G23"].number_format = money_fmt
    ws["G23"].font = bold_font

    # Payment conditions
    ws["A24"] = "CONDICIONES DE PAGO:"
    ws["A24"].font = Font(bold=True, size=12)
    ws["A25"] = pres["condiciones_pago"] or "50% adelanto-50% antes de la entrega del trabajo."
    ws["A25"].font = Font(size=12)
    ws["A26"] = f"Los presupuestos caducan a los {pres['dias_validez'] or 15} días"
    ws["A26"].font = Font(size=12)
    ws["A27"] = f"Pago mediante transferencia bancaria: CC: {empresa_cuenta}"
    ws["A27"].font = Font(bold=True, size=12)

    # Save
    base_output = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output")
    safe_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_"
                        for c in pres["cliente_nombre"]).strip()
    output_dir = os.path.join(base_output, safe_name)
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Presupuesto_{pres['numero_presupuesto']}_{safe_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    # If file locked, use timestamped name
    if os.path.exists(filepath):
        try:
            with open(filepath, "ab") as f:
                pass
        except PermissionError:
            from datetime import datetime as dt
            ts = dt.now().strftime("%H%M%S")
            filepath = os.path.join(output_dir, f"Presupuesto_{pres['numero_presupuesto']}_{safe_name}_{ts}.xlsx")
    wb.save(filepath)
    return filepath
